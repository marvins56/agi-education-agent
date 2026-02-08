"""Spreadsheet parser for XLSX, XLS, and CSV files.

Uses openpyxl for .xlsx files and pandas for .csv and .xls files.
Outputs sheet content as markdown tables.
"""

import logging
import pathlib

import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)

MAX_ROWS = 500


def _dataframe_to_markdown(df: pd.DataFrame, sheet_name: str = "") -> str:
    """Convert a pandas DataFrame to a markdown table string.

    If the DataFrame has more than MAX_ROWS rows, it is truncated
    with a note indicating how many rows were omitted.
    """
    if df.empty:
        return ""

    total_rows = len(df)
    truncated = total_rows > MAX_ROWS
    if truncated:
        df = df.head(MAX_ROWS)

    # Build header
    headers = [str(col).replace("|", "\\|") for col in df.columns]
    header_line = "| " + " | ".join(headers) + " |"
    separator_line = "| " + " | ".join("---" for _ in headers) + " |"

    lines = [header_line, separator_line]

    # Build data rows
    for _, row in df.iterrows():
        cells = []
        for val in row:
            cell_text = str(val).replace("|", "\\|") if pd.notna(val) else ""
            cells.append(cell_text)
        lines.append("| " + " | ".join(cells) + " |")

    result = "\n".join(lines)

    if truncated:
        remaining = total_rows - MAX_ROWS
        result += f"\n\n[...{remaining} more rows]"

    return result


class SpreadsheetParser:
    """Extract text and metadata from XLSX, XLS, and CSV files."""

    def parse(self, file_path: str) -> str:
        """Extract all content from a spreadsheet file as markdown tables."""
        path = pathlib.Path(file_path)
        suffix = path.suffix.lower()

        try:
            if suffix == ".xlsx":
                return self._parse_xlsx(file_path)
            elif suffix == ".csv":
                return self._parse_csv(file_path)
            elif suffix == ".xls":
                return self._parse_xls(file_path)
            else:
                logger.error("Unsupported spreadsheet format: %s", suffix)
                return ""
        except Exception as exc:
            logger.error("Failed to parse spreadsheet %s: %s", file_path, exc)
            return ""

    def parse_with_metadata(self, file_path: str) -> dict:
        """Extract content and metadata from a spreadsheet file."""
        path = pathlib.Path(file_path)
        suffix = path.suffix.lower()

        try:
            if suffix == ".xlsx":
                return self._parse_xlsx_with_metadata(file_path)
            elif suffix == ".csv":
                return self._parse_csv_with_metadata(file_path)
            elif suffix == ".xls":
                return self._parse_xls_with_metadata(file_path)
            else:
                logger.error("Unsupported spreadsheet format: %s", suffix)
                return {"text": "", "pages": 0, "metadata": {}}
        except Exception as exc:
            logger.error("Failed to parse spreadsheet %s: %s", file_path, exc)
            return {"text": "", "pages": 0, "metadata": {}}

    # ---- XLSX (openpyxl) ----

    def _parse_xlsx(self, file_path: str) -> str:
        """Parse an XLSX file using openpyxl, return markdown tables."""
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception as exc:
            logger.error("Failed to open XLSX file %s: %s", file_path, exc)
            return ""

        parts = []
        try:
            for sheet_name in wb.sheetnames:
                try:
                    ws = wb[sheet_name]
                    md = self._worksheet_to_markdown(ws, sheet_name)
                    if md:
                        parts.append(md)
                except Exception as exc:
                    logger.debug("Error processing sheet '%s': %s", sheet_name, exc)
                    continue
        finally:
            wb.close()

        return "\n\n".join(parts)

    def _parse_xlsx_with_metadata(self, file_path: str) -> dict:
        """Parse an XLSX file and return text with metadata."""
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception as exc:
            logger.error("Failed to open XLSX file %s: %s", file_path, exc)
            return {"text": "", "pages": 0, "metadata": {}}

        parts = []
        total_rows = 0
        total_columns = 0

        try:
            for sheet_name in wb.sheetnames:
                try:
                    ws = wb[sheet_name]
                    md = self._worksheet_to_markdown(ws, sheet_name)
                    if md:
                        parts.append(md)

                    # Count rows and columns
                    if ws.max_row:
                        total_rows += ws.max_row
                    if ws.max_column:
                        total_columns = max(total_columns, ws.max_column)
                except Exception as exc:
                    logger.debug("Error processing sheet '%s': %s", sheet_name, exc)
                    continue

            sheet_count = len(wb.sheetnames)
        finally:
            wb.close()

        return {
            "text": "\n\n".join(parts),
            "pages": sheet_count,
            "metadata": {
                "sheet_count": sheet_count,
                "total_rows": total_rows,
                "total_columns": total_columns,
            },
        }

    def _worksheet_to_markdown(self, ws, sheet_name: str) -> str:
        """Convert an openpyxl worksheet to a markdown table."""
        rows_data = []
        row_count = 0

        for row in ws.iter_rows(values_only=True):
            row_count += 1
            if row_count > MAX_ROWS + 1:  # +1 for header row
                break
            cells = []
            for cell in row:
                if cell is None:
                    cells.append("")
                else:
                    cells.append(str(cell).replace("|", "\\|"))
            rows_data.append(cells)

        if not rows_data:
            return ""

        # Normalize column count to max across all rows
        max_cols = max(len(r) for r in rows_data)
        for r in rows_data:
            while len(r) < max_cols:
                r.append("")

        lines = [f"### {sheet_name}"]

        # Header row
        lines.append("| " + " | ".join(rows_data[0]) + " |")
        lines.append("| " + " | ".join("---" for _ in rows_data[0]) + " |")

        # Data rows
        for row in rows_data[1:]:
            lines.append("| " + " | ".join(row) + " |")

        total_actual_rows = ws.max_row if ws.max_row else row_count
        if total_actual_rows > MAX_ROWS + 1:
            remaining = total_actual_rows - MAX_ROWS - 1
            lines.append(f"\n[...{remaining} more rows]")

        return "\n".join(lines)

    # ---- CSV (pandas) ----

    def _parse_csv(self, file_path: str) -> str:
        """Parse a CSV file using pandas, return as markdown table."""
        try:
            df = pd.read_csv(file_path, on_bad_lines="skip")
        except Exception as exc:
            logger.error("Failed to read CSV file %s: %s", file_path, exc)
            return ""

        return _dataframe_to_markdown(df)

    def _parse_csv_with_metadata(self, file_path: str) -> dict:
        """Parse a CSV file and return text with metadata."""
        try:
            df = pd.read_csv(file_path, on_bad_lines="skip")
        except Exception as exc:
            logger.error("Failed to read CSV file %s: %s", file_path, exc)
            return {"text": "", "pages": 0, "metadata": {}}

        text = _dataframe_to_markdown(df)

        return {
            "text": text,
            "pages": 1,
            "metadata": {
                "sheet_count": 1,
                "total_rows": len(df),
                "total_columns": len(df.columns),
            },
        }

    # ---- XLS (pandas) ----

    def _parse_xls(self, file_path: str) -> str:
        """Parse an XLS file using pandas, return as markdown tables."""
        try:
            sheets = pd.read_excel(file_path, sheet_name=None, engine="xlrd")
        except Exception as exc:
            logger.error("Failed to read XLS file %s: %s", file_path, exc)
            return ""

        parts = []
        for sheet_name, df in sheets.items():
            try:
                md = _dataframe_to_markdown(df, sheet_name)
                if md:
                    parts.append(f"### {sheet_name}\n\n{md}")
            except Exception as exc:
                logger.debug("Error processing XLS sheet '%s': %s", sheet_name, exc)
                continue

        return "\n\n".join(parts)

    def _parse_xls_with_metadata(self, file_path: str) -> dict:
        """Parse an XLS file and return text with metadata."""
        try:
            sheets = pd.read_excel(file_path, sheet_name=None, engine="xlrd")
        except Exception as exc:
            logger.error("Failed to read XLS file %s: %s", file_path, exc)
            return {"text": "", "pages": 0, "metadata": {}}

        parts = []
        total_rows = 0
        total_columns = 0

        for sheet_name, df in sheets.items():
            try:
                md = _dataframe_to_markdown(df, sheet_name)
                if md:
                    parts.append(f"### {sheet_name}\n\n{md}")
                total_rows += len(df)
                total_columns = max(total_columns, len(df.columns))
            except Exception as exc:
                logger.debug("Error processing XLS sheet '%s': %s", sheet_name, exc)
                continue

        return {
            "text": "\n\n".join(parts),
            "pages": len(sheets),
            "metadata": {
                "sheet_count": len(sheets),
                "total_rows": total_rows,
                "total_columns": total_columns,
            },
        }
